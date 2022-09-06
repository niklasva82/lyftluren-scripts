import openpyxl as op
import sys
from datetime import datetime, timedelta
from dateutil import parser
import re
import functools
import yaml

time_pattern = re.compile(r'(?P<hours>\d+):(?P<minutes>\d+):(?P<seconds>\d+)')

def parse_time(time_str):
    parts = time_pattern.match(time_str)
    if not parts:
        return
    parts = parts.groupdict()
    time_params = {}
    for name, param in parts.items():
        if param:
            time_params[name] = int(param)
    return timedelta(**time_params)

filename = sys.argv[1]

data = op.load_workbook(filename)

data = data.active

if len(sys.argv) > 2:
    start_time = sys.argv[2]
else:
    start_time = input("Välj början på tidsintervall (ex. 2022-08-25 12:00): ")

start_time = parser.parse(start_time)

if len(sys.argv) > 3:
    end_time = sys.argv[3]
else:
    end_time = input("Välj slut på tidsintervall (ex. 2022-08-25 12:00): ")

if len(sys.argv) > 4:
    caller = sys.argv[4]
else:
    caller = input("Välj ringare (ex. \"Ringare 2 Stockholm\" eller \"*\"): ")

end_time = parser.parse(end_time)

result = {
}

call_lengths = []
call_times = []
handling_times = []
min_time = end_time
max_time = start_time
locations = {}
caller_times = {}

not_calls = {
    'Positiv': 0,
    'Negativ': 0,
    'Vet ej': 0,
}

time_not_calls = {
    'Positiv': timedelta(seconds=0),
    'Negativ': timedelta(seconds=0),
    'Vet ej': timedelta(seconds=0),
}

for row in data.iter_rows(1, data.max_row):
    if row[0].value != 'Stockholm':
        continue
    position = row[9].value
    call_datetime = parser.parse(row[10].value)
    call_datetime_end = call_datetime + parse_time(row[8].value)
    if call_datetime < start_time or call_datetime_end > end_time:
        continue
    if caller != '*' and row[11].value != caller:
        continue

    if call_datetime < min_time:
        min_time = call_datetime
    if call_datetime > max_time:
        max_time = call_datetime

    if position in result:
        result[position] += 1
    else:
        result[position] = 1
    if position in ('Positiv', 'Negativ', 'Vet ej'):
        call_lengths.append(parse_time(row[8].value))
        if parse_time(row[8].value) < timedelta(seconds=3):
            not_calls[position] += 1
            time_not_calls[position] += parse_time(row[8].value)
    date = str(call_datetime.date())
    #if call_datetime > parser.parse("2022-08-23"):
    #    import pdb; pdb.set_trace()
    if row[11].value in caller_times:
        if date in caller_times[row[11].value]:
            #if call_datetime < caller_times[row[11].value][date][-1]['begin']:
            #    caller_times[row[11].value][date][-1]['begin'] = call_datetime
            if call_datetime > caller_times[row[11].value][date][-1]['end']:
                if call_datetime > caller_times[row[11].value][date][-1]['end'] + timedelta(minutes=20):
                    caller_times[row[11].value][date].append({
                        'begin': call_datetime,
                        'end': call_datetime_end,
                    })
                else:
                    caller_times[row[11].value][date][-1]['end'] = call_datetime_end
        else:
            caller_times[row[11].value][date] = [{
                'begin': call_datetime,
                'end': call_datetime_end,
            }]
    else:
        caller_times[row[11].value] = {}
        caller_times[row[11].value][date] = [{
            'begin': call_datetime,
            'end': call_datetime_end,
        }]

    call_times.append(parse_time(row[12].value))
    handling_times.append(parse_time(row[13].value))
    location = row[7].value
    if location in locations:
        locations[location] += 1
    else:
        locations[location] = 1


positiv = result.get('Positiv') or 0
negativ = result.get('Negativ') or 0
vet_ej = result.get('Vet ej') or 0
ej_vald = result.get('Ej vald') or 0
telefonsvarare = result.get('Telefonsvarare') or 0
upptaget = result.get('Upptaget') or 0
fortsatt = result.get('Fortsatt samtal') or 0
svarade = positiv + negativ + vet_ej

tot = positiv + negativ + vet_ej + ej_vald + telefonsvarare + upptaget + fortsatt

print("Uppringingar: ", str(tot))
print("----------------")
print("Svarade: %d %.2f" % (svarade, (100*svarade/tot)), "%")
print("----------------")
print("Positiv: %d %.2f" % (positiv, (100*positiv/svarade)), "%")
print("Negativ: %d %.2f" % (negativ, (100*negativ/svarade)), "%")
print("Vet ej: %d %.2f" % (vet_ej, (100*vet_ej/svarade)), "%")
print("----------------")
print("Telefonsvarare: %d %.2f" % (telefonsvarare, (100*telefonsvarare/tot)), "%")
print("Ej vald: %d %.2f" % (ej_vald, (100*ej_vald/tot)), "%")
print("Upptaget: %d %.2f" % (upptaget, (100*upptaget/tot)), "%")
print("Fortsatt samtal: %d %.2f" % (fortsatt, (100*fortsatt/tot)), "%")
print("----------------")

total_call_length = timedelta(seconds=0)
max_call_length = timedelta(seconds=0)
for le in call_lengths:
    total_call_length += le
    if le > max_call_length:
        max_call_length = le

total_call_time = timedelta(seconds=0)
max_call_time = timedelta(seconds=0)
for le in call_times:
    total_call_time += le
    if le > max_call_length:
        max_call_time = le

total_call_time = timedelta(seconds=0)
max_call_time = timedelta(seconds=0)
for le in call_times:
    total_call_time += le
    if le > max_call_time:
        max_call_time = le

total_handling_time = timedelta(seconds=0)
max_handling_time = timedelta(seconds=0)
for le in handling_times:
    total_handling_time += le
    if le > max_handling_time:
        max_handling_time = le

print("Total samtalslängd (exklusive telefonsvarare):", str(total_call_length))
print("Genomsnittlig samtalslängd: ", str(total_call_length/len(call_lengths)))
print("Maximal samtalslängd: ", str(max_call_length))
print("=====================")
print("För korta posititiva samtal (mindre än 3 sekunder)", not_calls['Positiv'])
print("För korta negativa samtal (mindre än 3 sekunder)", not_calls['Negativ'])
print("För korta vet ej samtal (mindre än 3 sekunder)", not_calls['Vet ej'])
print("=====================")
if not_calls['Positiv']:
    print("För korta posititiva samtal (mindre än 3 sekunder)", str(time_not_calls['Positiv']/not_calls['Positiv']))
if not_calls['Negativ']:
    print("För korta negativa samtal (mindre än 3 sekunder)", str(time_not_calls['Negativ']/not_calls['Negativ']))
if not_calls['Vet ej']:
    print("För korta vet ej samtal (mindre än 3 sekunder)", str(time_not_calls['Vet ej']/not_calls['Vet ej']))
print("=====================")
print("Total ringtidtid", str(total_call_time))
print("Max ringtid", str(max_call_time))
print("=====================")
print("Total hanteringstid", str(total_handling_time))
print("Max hanteringstid", str(max_handling_time))
print("=====================")

total_hours = timedelta(seconds=0)
for c in caller_times:
    for dt in caller_times[c]:
        for i in range(len(caller_times[c][dt])):
            total_hours += caller_times[c][dt][i]['end'] - caller_times[c][dt][i]['begin']
total_hours = total_hours.total_seconds()/3600

print("Ringtid (exkl. pauser mer än 20 min):", total_hours, "hours")
print("Uppringningar per ringtimme: ", str(tot/total_hours))
print("Samtal per ringtimme: ", str(svarade/total_hours))
print("Platser ringda:")
print("=====================")


for l in locations:
    print(l, ": ", locations[l])
